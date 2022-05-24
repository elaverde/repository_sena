import sys
import os
import openpyxl
from multipledispatch import dispatch
from openpyxl.styles import Alignment 
class Excel:
    def __init__(self):
        self.excel = None
        self.sheet = None
        self.row = None
        self.col = None
        self.value = None
        self.cell = None

    def created_file(self):
        self.excel = openpyxl.Workbook()
        self.sheet = self.excel.get_sheet_by_name("Sheet")
     

    def set_width_column(self,column, width):
        self.sheet.column_dimensions[column].width = width

    def save_file(self,filepath):
        self.excel.save(filepath)

    def read_file(self,url):
        self.excel = openpyxl.load_workbook(url)
    
    def read_sheet(self,sheet):
        self.sheet = self.excel.get_sheet_by_name(sheet)
 
    
    def read_cell(self,cell):
        self.cell = cell
        self.value = self.value = self.sheet[cell].value 
        return self.value

    def write_cell(self,cell,value):
        self.cell = cell
        self.value = value
        self.sheet[cell].value = self.value

    def merge_cell(self,start,finish):
        self.sheet.merge_cells(start+':'+finish)  

    def center_cell(self,cell):
        cell = self.sheet[cell]
        #cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.alignment = Alignment(vertical='center')